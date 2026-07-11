"""Report exports: Excel schedule and CSV."""
from __future__ import annotations

import csv
import io

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from ..database import get_db
from .. import models, services

router = APIRouter(prefix="/api/projects/{project_id}/report", tags=["reports"])

COLUMNS = [
    ("day", "Day"), ("the_date", "Date"), ("julian_day", "Julian"),
    ("stage", "Stage"), ("eto", "ETo (mm)"), ("kc", "Kc"), ("etc", "ETc (mm)"),
    ("rainfall", "Rain (mm)"), ("pe", "Pe (mm)"), ("lr", "LR"), ("iwr", "IWR (mm)"),
    ("root_depth", "Zr (m)"), ("taw", "TAW (mm)"), ("raw", "RAW (mm)"),
    ("p", "p"), ("ds_begin", "Ds begin"), ("daily_water_use", "Daily use"),
    ("ds_end", "Ds end"), ("irrigate", "Irrigate"), ("dn", "Dn (mm)"),
    ("dg", "Dg (mm)"), ("volume_m3", "Volume (m3)"), ("runtime_hours", "Runtime (h)"),
]


def _proj(db, project_id):
    proj = db.get(models.Project, project_id)
    if not proj:
        raise HTTPException(404, "Project not found")
    if not proj.climate:
        raise HTTPException(400, "Project has no climate data.")
    return proj


@router.get("/csv")
def report_csv(project_id: int, db: Session = Depends(get_db)):
    proj = _proj(db, project_id)
    res = services.run_schedule(proj)
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow([h for _, h in COLUMNS])
    for d in res.daily:
        row = d.to_dict()
        w.writerow([row.get(k) for k, _ in COLUMNS])
    buf.seek(0)
    fname = f"schedule_{proj.project_name.replace(' ', '_')}.csv"
    return StreamingResponse(
        iter([buf.getvalue()]), media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'})


@router.get("/excel")
def report_excel(project_id: int, db: Session = Depends(get_db)):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    proj = _proj(db, project_id)
    res = services.run_schedule(proj)
    wb = openpyxl.Workbook()

    # --- Summary sheet ---
    ws = wb.active
    ws.title = "Summary"
    crop, soil, system = proj.crop, proj.soil, proj.system
    info = [
        ("Smart Irrigation Scheduling Platform — FAO-56", ""),
        ("Project", proj.project_name), ("Farm", proj.farm_name),
        ("Field", proj.field_name),
        ("Location", f"{proj.city}, {proj.region}, {proj.country}"),
        ("Latitude / Longitude", f"{proj.latitude} / {proj.longitude}"),
        ("Elevation (m)", proj.elevation),
        ("Area", f"{proj.area_value} {proj.area_unit} ({proj.area_m2:.0f} m²)"),
        ("Planting date", proj.planting_date.isoformat()),
        ("Crop", f"{crop.name_en} ({crop.name_ar})"),
        ("Soil", f"{soil.name_en} (FC={soil.theta_fc}, WP={soil.theta_wp})"),
        ("Irrigation system", f"{system.name_en} @ {proj.efficiency_pct}%"),
        ("Water salinity ECw / ECe", f"{proj.ecw} / {proj.ece} dS/m"),
        ("", ""),
        ("Season length (days)", res.summary.days),
        ("Number of irrigations", res.summary.n_irrigations),
        ("Total ETo (mm)", res.summary.total_eto),
        ("Total ETc (mm)", res.summary.total_etc),
        ("Effective rainfall (mm)", res.summary.total_effective_rain),
        ("Total net depth (mm)", res.summary.total_net_depth),
        ("Total gross depth (mm)", res.summary.total_gross_depth),
        ("Gross water volume (m³)", res.summary.total_gross_volume_m3),
        ("Total runtime (hours)", res.summary.total_runtime_hours),
    ]
    ws["A1"].font = Font(bold=True, size=14)
    for i, (k, v) in enumerate(info, start=1):
        ws.cell(i, 1, k).font = Font(bold=True)
        ws.cell(i, 2, v)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 40

    # --- Daily schedule sheet ---
    ds = wb.create_sheet("Daily Schedule")
    header_fill = PatternFill("solid", fgColor="1B5E20")
    for c, (_, h) in enumerate(COLUMNS, start=1):
        cell = ds.cell(1, c, h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    irr_fill = PatternFill("solid", fgColor="C8E6C9")
    for r, d in enumerate(res.daily, start=2):
        row = d.to_dict()
        for c, (k, _) in enumerate(COLUMNS, start=1):
            cell = ds.cell(r, c, row.get(k))
            if row.get("irrigate"):
                cell.fill = irr_fill
    ds.freeze_panes = "A2"

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    fname = f"irrigation_report_{proj.project_name.replace(' ', '_')}.xlsx"
    return StreamingResponse(
        out, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'})


@router.get("/pdf")
def report_pdf(project_id: int, db: Session = Depends(get_db)):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer, Table,
                                    TableStyle)

    proj = _proj(db, project_id)
    res = services.run_schedule(proj)
    s, crop, soil, system = res.summary, proj.crop, proj.soil, proj.system

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=18 * mm, bottomMargin=15 * mm,
                            leftMargin=15 * mm, rightMargin=15 * mm,
                            title=f"Irrigation Report — {proj.project_name}")
    styles = getSampleStyleSheet()
    green = colors.HexColor("#1b5e20")
    h1 = ParagraphStyle("h1", parent=styles["Title"], textColor=green, fontSize=17)
    h2 = ParagraphStyle("h2", parent=styles["Heading2"], textColor=green, fontSize=12)
    small = ParagraphStyle("small", parent=styles["Normal"], fontSize=8,
                           textColor=colors.grey)
    story = []

    story.append(Paragraph("Smart Irrigation Scheduling Platform — FAO-56", h1))
    story.append(Paragraph("Crop Water Requirement & Irrigation Schedule Report", styles["Normal"]))
    story.append(Spacer(1, 8 * mm))

    def kv_table(title, rows):
        story.append(Paragraph(title, h2))
        t = Table([[k, str(v)] for k, v in rows], colWidths=[70 * mm, 100 * mm])
        t.setStyle(TableStyle([
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("TEXTCOLOR", (0, 0), (0, -1), colors.HexColor("#5b6770")),
            ("LINEBELOW", (0, 0), (-1, -1), 0.3, colors.HexColor("#e2e8e4")),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
        ]))
        story.append(t)
        story.append(Spacer(1, 5 * mm))

    kv_table("Project Information", [
        ("Project", proj.project_name), ("Farm / Field", f"{proj.farm_name} / {proj.field_name}"),
        ("Location", f"{proj.city}, {proj.region}, {proj.country}"),
        ("Latitude / Longitude", f"{proj.latitude} / {proj.longitude}"),
        ("Elevation", f"{proj.elevation} m"),
        ("Area", f"{proj.area_value} {proj.area_unit} ({proj.area_m2:.0f} m²)"),
        ("Planting date", proj.planting_date.isoformat()),
        ("Crop", f"{crop.name_en} ({crop.name_ar}) — {crop.scientific_name}"),
        ("Soil", f"{soil.name_en} — FC {soil.theta_fc}, WP {soil.theta_wp}"),
        ("Irrigation system", f"{system.name_en} @ {proj.efficiency_pct}%"),
        ("Water salinity ECw / ECe", f"{proj.ecw} / {proj.ece} dS/m"),
    ])

    kv_table("Crop Parameters (FAO-56)", [
        ("Stage lengths (Lini/Ldev/Lmid/Llate)", f"{crop.l_ini}/{crop.l_dev}/{crop.l_mid}/{crop.l_late} d"),
        ("Kc (ini / mid / end)", f"{crop.kc_ini} / {crop.kc_mid} / {crop.kc_end}"),
        ("Root depth Zr", f"{crop.zr_min}–{crop.zr_max} m"),
        ("Depletion fraction p", f"{crop.p}"), ("Yield response Ky", f"{crop.ky}"),
    ])

    kv_table("Water Requirement Summary", [
        ("Season length", f"{s.days} days"), ("Number of irrigations", f"{s.n_irrigations}"),
        ("Total ETo", f"{s.total_eto} mm"), ("Total ETc (crop water requirement)", f"{s.total_etc} mm"),
        ("Effective rainfall", f"{s.total_effective_rain} mm"),
        ("Net irrigation depth", f"{s.total_net_depth} mm"),
        ("Gross irrigation depth", f"{s.total_gross_depth} mm"),
        ("Gross water volume", f"{s.total_gross_volume_m3} m³"),
        ("Total pump runtime", f"{s.total_runtime_hours} h"),
    ])

    story.append(Paragraph("FAO-56 Methodology", h2))
    for f in ["ETo = FAO-56 Penman-Monteith (Eq. 6)", "ETc = Kc × ETo",
              "TAW = 1000·(θFC − θWP)·Zr,  RAW = p·TAW",
              "Irrigate when Dr ≥ RAW,  Dg = Dn / ((1 − LR)·Ea)"]:
        story.append(Paragraph(f, styles["Code"]))
    story.append(Spacer(1, 6 * mm))
    story.append(Paragraph(
        "Reference ET uses canonical FAO-56 constants (γ = 0.000665·P, albedo 0.23, "
        "σ = 4.903×10⁻⁹). All default values are user-editable.", small))

    doc.build(story)
    buf.seek(0)
    fname = f"irrigation_report_{proj.project_name.replace(' ', '_')}.pdf"
    return StreamingResponse(
        buf, media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'})
